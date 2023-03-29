#!/usr/bin/env python3

# -*- coding: utf-8 -*-
"""
Created on Wed Nov 24 16:10:08 2021

@author: darshika_verma
"""

import openpyxl
wb = openpyxl.Workbook()

#Creating workbook
wb.save("First Workbook.xlsx")

#Opening an existing workbook
wb = openpyxl.load_workbook("First Workbook.xlsx")
for sheet in wb:
    print(sheet.title)  #Name of Sheet in Excel
    
    
"""
Creating a Workbook object with OpenPyXL
Adding, deleting and renaming sheets
Saving the workbook to disk
"""
from openpyxl import Workbook

def main():
    #Create the Workbook object
    wb = Workbook()

    #Create a new sheet
    ws = wb.create_sheet("A sheet", 0) # insert at first position

    # Change the name of the sheet
    ws.title = "Hello World!"

    #Create another sheet
    ws2 = wb.create_sheet("Sheet nr 2") # Insert the new sheet at the end

    # Check what sheets exists in the workbook
    print("Sheets in workbook:")
    for sheet in wb:
        print(sheet.title)
    print("-"*20)    
    # Delete sheet "Sheet"
    wb.remove(wb["Sheet"])
    #del wb["Sheet"] also works

    # Check what sheets are left
    print("Sheets in workbook after deletion:")
    for sheet in wb:
        print(sheet.title)

    # Save the workbook to disc
    wb.save('2.3_Hello_sheets.xlsx')
    print("Exiting main()")

if __name__ == "__main__":
    main()

############################################################################################
"""
Open an existing workbook, copy one sheet
Copy data from one sheet to another
"""
from openpyxl import load_workbook

def main():
    #Create the Workbook object
    wb = load_workbook("2.3_Hello_sheets.xlsx")

    #Copy sheets
    source = wb["Sheet nr 2"]
    new_sheet = wb.copy_worksheet(source)
    new_sheet.title = "Copy of Sheet nr 2"

    # Check what sheets exists in the workbook
    print("Sheets in workbook:")
    for sheet in wb:
        print(sheet.title)

    wb.save("2.4_Hello_copies.xlsx")

if __name__ == "__main__":
    main()
    
############################################################################################
"""
Get sheets by index or name
"""
from openpyxl import load_workbook

def main():
    # Open the Hello_copies.xlsx
    wb = load_workbook("2.4_Hello_copies.xlsx")
    # Check what indices the sheets have
    for sheet in wb:
        print("{} has the index {}".format(sheet.title, wb.index(sheet)))
    print("-"*20)
    # Get by index method 1
    ws1 = wb.worksheets[0]
    ws2 = wb.worksheets[1]
    # Get by index method 2
    worksheets = wb.sheetnames  
    ws3 = wb[worksheets[2]]
    print("The first sheet has the title", ws1.title, "\nThe second sheet has the title:", ws2.title \
          , "\nThe third sheet has the title:", ws3.title)
    # Get a sheet by name
    ws2 = wb["Hello World!"]
    
if __name__ == "__main__":
    main()
##########################################################################################
"""
Summary functions of chapter 2
"""
from openpyxl import load_workbook
from openpyxl import Workbook

def save_wb(wb, filename):
    # Save a workbook
    wb.save(filename)

def open_wb(filename):
    # Returns an opened workbook
    return load_workbook(filename)

def create_sheets(wb, sheet_name_list):
    # Adds the sheets in the sheet_name_list to the workbook
    for sheet_name in sheet_name_list:
        wb.create_sheet(sheet_name)

def delete_sheet_by_name(wb, sheet_name):
    wb.remove(wb[sheet_name])

def copy_sheet(wb, source_sheet_name, new_sheet_name=""):
    new_sheet_name = "Copy of " + source_sheet_name if new_sheet_name == "" else new_sheet_name
    #Copy sheets
    source = wb[source_sheet_name]
    new_sheet = wb.copy_worksheet(source)
    new_sheet.title = new_sheet_name

def get_sheet_name_and_index_from_wb(wb):
    # Return a dictionary holding the indexes and names of the sheets
    sheet_name_index_dict = {}
    for index, sheet in enumerate(wb):
        sheet_name_index_dict[index] = sheet.title
    return sheet_name_index_dict

def get_sheet_by_index(wb, index):
    # Returns a sheet at the provided index
    try:
        return wb.worksheets[index]
    except IndexError:
        print("No sheet exists with index", index)

if __name__ == "__main__":
    wb = Workbook()
    create_sheets(wb, ["A sheet", "Another sheet", "Yet another sheet"])
    delete_sheet_by_name(wb, "Yet another sheet")
    copy_sheet(wb, "A sheet")
    copy_sheet(wb, "A sheet", "Fresh copy")
    print(get_sheet_by_index(wb, 3).title)
    #print(get_sheet_by_index(wb, 56).title)
    print("-"*60)
    print(get_sheet_name_and_index_from_wb(wb))
    
###################################################################################################

"""
Explore relative and absolute referencing of cells
"""
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string


def save_wb(wb, filename):
    # Save a workbook
    wb.save(filename)

def create_sheets(wb, sheet_name_list):
    # Adds the sheets in the sheet_name_list to the workbook
    for sheet_name in sheet_name_list:
        wb.create_sheet(sheet_name)

if __name__ == "__main__":
    # Create a workbook
    filename = "Absolute_relative.xlsx"
    wb = Workbook()
    create_sheets(wb, ["Sheet2", "Sheet3", "Sheet4"])

    # Input some values in the first sheet with Relative references
    ws1 = wb.worksheets[0]
    ws1["A1"] = 550
    ws1["C3"] = "Caravan"

    print("get_column_letter for index 3 is:", get_column_letter(3))
    print("column_index_from_string for letter C is:", column_index_from_string("C"))

    # Input values in Sheet2 using Absolute references
    ws2 = wb["Sheet2"]
    ws2.cell(row=4, column=2, value= 10)    # Cell B4
    ws2.cell(1, 2).value = "Train"          # Cell B1

    # Save the wb
    save_wb(wb, filename)
    
###############################################################################################

"""
Cell offset
"""
from openpyxl import load_workbook
from openpyxl import Workbook

def save_wb(wb, filename):
    # Save a workbook
    wb.save(filename)

def create_sheets(wb, sheet_name_list):
    # Adds the sheets in the sheet_name_list to the workbook
    for sheet_name in sheet_name_list:
        wb.create_sheet(sheet_name)

if __name__ == "__main__":
    # Create a workbook an sheets
    filename = "Absolute_relative.xlsx"
    wb = Workbook()
    create_sheets(wb, ["Sheet2", "Sheet3", "Sheet4"])
    ws1 = wb["Sheet"]

    # Set value of cell B1 to Train
    ws1.cell(1, 2).value = "Train"
    # Set value of cell C1 to Train cart 
    ws1.cell(1, 2).offset(0, 1).value = "Train cart"

    # Create cell objects for easier references
    mother_cell = ws1.cell(3,3) # Cell C3
    child_cell = mother_cell.offset(1,0) # Cell C4
    mother_cell.value = "Mother"
    child_cell.value = "Child"
    
    # Save the wb
    save_wb(wb, filename)
